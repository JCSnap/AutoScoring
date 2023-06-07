import pandas as pd
from openpyxl import load_workbook

# Open the text file and read it
with open('results.txt', 'r') as file:
    data = file.read()

# Split the text file into separate blocks
blocks = data.split('Question: ')

# Define the columns names
columns = ["Question", "Model Answer", "Student Answer", "OpenAI Justification",
           "OpenAI Score", "Original Score", "Time taken for OpenAI API", "Time taken for Original API"]

# Initialize an empty list to store the data
rows = []

# Loop over each block
for block in blocks[1:]:  # We start from the second block because the first block is empty
    block = 'Question: ' + block  # Add back the 'Question: ' prefix we removed earlier

    # Split the block into lines
    lines = block.split('\n')

    # Initialize an empty dictionary to store the data for this block
    row = {}

    # Loop over each line in the block
    for line in lines:
        # Find the first colon in the line
        index = line.find(':')

        # Get the field and value
        field = line[:index].strip()
        value = line[index+1:].strip()

        # Remove any trailing 'seconds' in the 'Time taken' fields
        if 'Time taken' in field:
            value = value.replace(' seconds', '')

        # Convert the numeric fields to float
        if field in ['OpenAI Score', 'Original Score', 'Time taken for OpenAI API', 'Time taken for Original API']:
            value = float(value)

        # Add the field and value to the row dictionary
        row[field] = value

    # Add the row dictionary to the rows list
    rows.append(row)

# Create a DataFrame from the rows list
df = pd.DataFrame(rows, columns=columns)

# Check if the Excel file already exists
try:
    # Try to load the existing workbook
    book = load_workbook('visualize.xls')

    # Get the active worksheet
    writer = pd.ExcelWriter('visualize.xls', engine='openpyxl')
    writer.book = book

    # Write the DataFrame to the worksheet, starting at the last row
    df.to_excel(writer, index=False, header=False,
                startrow=writer.sheets['Sheet1'].max_row)

    writer.save()

except FileNotFoundError:
    # If the file does not exist, just write the DataFrame to a new file
    df.to_excel('visualize.xls', index=False)
